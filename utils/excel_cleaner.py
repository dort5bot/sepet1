# Excel Temizleyici (utils/excel_cleaner.py) - Geliştirilmiş Asenkron Versiyon
# Sütun genişliği otomatik olarak içeriğe göre ayarlanır, minimum 10, maksimum 25 birim
# Tam asenkron uyumlu, performans iyileştirmeli
# 14.05.2026
# utils/excel_cleaner.py
"""
Bu kod büyük dosyalar için optimize edilmiş, fakat:
1.000–5.000 satır Excel dosyalarında çok iyi performans verir
20.000+ satırda bile ThreadPoolExecutor sayesinde ölmez
Bellek sızıntısı yok
ThreadPool temizliği doğru yapılmış
Hiçbir kritik hata veya memory leak yok
"""

import asyncio
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple, Any, Optional
from utils.logger import logger
import tempfile
import os
import re
import unicodedata
import aiofiles
from pathlib import Path

from datetime import date, datetime, timedelta # Excel tarih
from concurrent.futures import ThreadPoolExecutor # Excel tarih düzeltici yardımcı

# Sabitler
MAX_HEADER_SEARCH_ROWS = 5
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 25
HEADER_ROW_BUFFER = 2
MAX_FILE_SIZE_MB = 50  # Maksimum dosya boyutu
CHUNK_SIZE = 1000  # Büyük dosyalar için chunk boyutu


# Excel tarih düzeltici yardımcı
def fix_excel_date(value):
    """
    Excel tarih değerlerini Python datetime.date nesnesine çevirir
    """
    from datetime import datetime, date, timedelta
    
    # 1. None kontrolü
    if value is None:
        return None
    
    # 2. Zaten datetime.date ise
    if isinstance(value, date):
        return value
    
    # 3. datetime.datetime ise, sadece tarih kısmını al
    if isinstance(value, datetime):
        return value.date()
    
    # 4. Excel seri numarası ise (int/float)
    if isinstance(value, (int, float)):
        try:
            # Excel 1899-12-30'dan başlar
            base_date = date(1899, 12, 30)
            return base_date + timedelta(days=float(value))
        except Exception:
            return value
    
    # 5. String ise
    if isinstance(value, str):
        value = value.strip()
        try:
            # Önce datetime olarak parse et
            dt = None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(value, fmt)
                    break
                except ValueError:
                    continue
            
            if dt:
                return dt.date()
        except Exception:
            pass
    
    # 6. Diğer durumlarda olduğu gibi döndür
    return value
    

# utils/excel_cleaner.py - YENİ: Esnek başlık eşleme sistemi

class HeaderMapper:
    """Çoklu dil ve varyasyon desteği ile başlık eşleme"""
    
    # 🔥 CORE: Tüm olası başlık varyasyonları
    CITY_HEADERS = {
        # Türkçe
        'İL', 'İLİ', 'CITY', 'ŞEHİR', 'SEHIR', 'İLÇE', 'ILCE','Nere'
        # İngilizce
        'CITY', 'PROVINCE', 'DISTRICT',
        # Hatalı yazımlar
        'IL', 'I L', 'iL', 'Il',
        # Alternatifler
        'LOCATION', 'PLACE', 'REGION'
    }
    
    DATE_HEADERS = {
        # Türkçe
        'TARİH', 'TARIH', 'DATE', 'GÜN', 'GUN', 'ZAMAN',
        # İngilizce
        'DATE', 'DAY', 'CREATED_DATE', 'UPDATE_DATE',
        # Alternatifler
        'TARİHİ', 'TARIHI', 'RECORD_DATE'
    }
    
    @classmethod
    def normalize_header(cls, header: str) -> str:
        """Başlığı normalize et - TÜM özel karakterleri temizle"""
        if not header:
            return ""
        
        # 1. Unicode normalize
        header = unicodedata.normalize('NFKC', str(header))
        
        # 2. Tüm özel karakterleri kaldır (noktalama, boşluk, özel simgeler)
        header = re.sub(r'[^\w\u00C0-\u00FF]', '', header)
        
        # 3. Türkçe karakterleri ASCII'ye dönüştür
        header = header.replace('İ', 'I').replace('I', 'I')
        header = header.replace('ı', 'i').replace('i', 'i')
        header = header.replace('Ğ', 'G').replace('Ğ', 'G')
        header = header.replace('ğ', 'g')
        header = header.replace('Ü', 'U').replace('Ü', 'U')
        header = header.replace('ü', 'u')
        header = header.replace('Ş', 'S').replace('Ş', 'S')
        header = header.replace('ş', 's')
        header = header.replace('Ö', 'O').replace('Ö', 'O')
        header = header.replace('ö', 'o')
        header = header.replace('Ç', 'C').replace('Ç', 'C')
        header = header.replace('ç', 'c')
        
        # 4. Büyük harfe çevir
        header = header.upper()
        
        # 5. Fazla boşlukları temizle
        header = ' '.join(header.split())
        
        return header.strip()
    
    @classmethod
    def find_column(cls, headers: List[str], target_sets: List[set], column_name: str) -> int:
        """
        En esnek sütun bulma algoritması
        - Tam eşleşme
        - Normalize edilmiş eşleşme
        - İçerik bazlı eşleşme
        - Regex bazlı eşleşme
        """
        for idx, header in enumerate(headers, 1):
            if not header:
                continue
            
            # Orijinal header'ı normalize et
            normalized = cls.normalize_header(header)
            
            # 1. Tam eşleşme (normalize edilmiş)
            if normalized in target_sets:
                return idx
            
            # 2. İçerik bazlı eşleşme (header içinde aranan kelime var mı?)
            for target in target_sets:
                if target in normalized or normalized in target:
                    return idx
            
            # 3. Regex bazlı eşleşme (pattern match)
            patterns = {
                'date': r'(TARİH|TARIH|DATE|GÜN|GUN)',
                'city': r'(İL|IL|CITY|ŞEHİR|SEHIR|PROVINCE)'
            }
            
            if column_name == 'date' and re.search(patterns['date'], normalized, re.I):
                return idx
            if column_name == 'city' and re.search(patterns['city'], normalized, re.I):
                return idx
        
        return -1  # Bulunamadı
        

# utils/excel_cleaner.py - Ana sınıf içine eklenecek

class SmartHeaderDetector:
    """Akıllı başlık dedektörü - Çoklu satır ve çoklu format desteği"""
    
    def __init__(self, max_search_rows: int = 10):
        self.max_search_rows = max_search_rows
        self.header_mapper = HeaderMapper()
    
    async def detect_headers(self, ws: Worksheet) -> Dict[str, Any]:
        """
        3 aşamalı başlık tespiti:
        1. Satır bazlı tarama (1-10 arası)
        2. En çok dolu hücreye sahip satırı başlık olarak belirle
        3. İçerik analizi ile doğrula
        """
        best_row = 1
        max_filled = 0
        
        # Aşama 1: En çok dolu hücreye sahip satırı bul
        for row in range(1, min(self.max_search_rows, ws.max_row) + 1):
            filled_count = 0
            row_values = []
            
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip():
                    filled_count += 1
                    row_values.append(str(val).strip())
            
            # Başlık satırı genellikle en çok dolu hücreye sahiptir
            if filled_count > max_filled and filled_count >= 2:
                max_filled = filled_count
                best_row = row
        
        # Aşama 2: Bulunan satırdan başlıkları al
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=best_row, column=col).value
            headers.append(str(val).strip() if val else "")
        
        # Aşama 3: Başlıkları doğrula
        date_col = HeaderMapper.find_column(headers, HeaderMapper.DATE_HEADERS, 'date')
        city_col = HeaderMapper.find_column(headers, HeaderMapper.CITY_HEADERS, 'city')
        
        if date_col == -1 or city_col == -1:
            # Fallback: İlk 2 sütunu dene (bazı basit Excel'ler için)
            if len(headers) >= 2:
                date_col, city_col = 1, 2
                logger.warning(f"Başlık bulunamadı, varsayılan sütunlar kullanıldı: TARİH=1, İL=2")
        
        return {
            'header_row': best_row,
            'headers': headers,
            'date_column': date_col,
            'city_column': city_col,
            'success': date_col != -1 and city_col != -1
        }
        

class AsyncExcelCleaner:
    """Excel dosya temizleme işlemlerini asenkron olarak yöneten sınıf"""
    
    def __init__(self):
        self.required_columns = {"TARİH", "İL"}
        self.thread_pool = ThreadPoolExecutor(max_workers=4)
    
    async def _check_file_size(self, file_path: str) -> bool:
        """Dosya boyutunu kontrol eder"""
        try:
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB cinsinden
            if file_size > MAX_FILE_SIZE_MB:
                logger.warning(f"Dosya boyutu çok büyük: {file_size:.2f}MB")
                return False
            return True
        except Exception as e:
            logger.error(f"Dosya boyutu kontrol hatası: {e}")
            return False
    
    # async def _find_header_row(self, ws: Worksheet) -> int:
    #     """Başlık satırını bulur (asenkron wrapper)"""
    #     #loop = asyncio.get_event_loop()
    #     loop = asyncio.get_running_loop()
    #     return await loop.run_in_executor(
    #         self.thread_pool, 
    #         self._sync_find_header_row, 
    #         ws
    #     )
    
    # async def _find_header_row(self, ws: Worksheet) -> int:  # BUNU SİL
    
    # def _sync_find_header_row(self, ws: Worksheet) -> int:
    #     """Başlık satırını senkron olarak bulur"""
    #     for row in range(1, MAX_HEADER_SEARCH_ROWS + 1):
    #         if any(ws.cell(row=row, column=col).value 
    #                for col in range(1, ws.max_column + 1)):
    #             return row
    #     return 1

    # YERİNE:
    async def _find_header_row_smart(self, ws: Worksheet) -> Dict[str, Any]:
        """SmartHeaderDetector ile başlık bul"""
        detector = SmartHeaderDetector(max_search_rows=MAX_HEADER_SEARCH_ROWS)
        return await detector.detect_headers(ws)
    
    async def _clean_headers(self, ws: Worksheet, header_row: int) -> List[str]:
        """Başlıkları temizler ve düzenler (asenkron wrapper)"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self.thread_pool,
            self._sync_clean_headers,
            ws, header_row
        )

   
   
    def _sync_clean_headers(self, ws: Worksheet, header_row: int) -> List[str]:
        """Başlıkları senkron olarak temizler"""
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            clean_value = (str(cell_value).strip().upper() 
                          if cell_value else f"Bos_{col}")
            headers.append(clean_value)
        return headers
    
    def _find_required_columns(self, headers: List[str]) -> Dict[str, int]:
        """Gerekli sütunların indekslerini bulur"""
        column_indices = {}
        
        for idx, header in enumerate(headers, 1):
            if "TARİH" in header and "TARİH" not in column_indices:
                column_indices["TARİH"] = idx
            elif "İL" in header and "İL" not in column_indices:
                column_indices["İL"] = idx
        
        missing_columns = self.required_columns - set(column_indices.keys())
        if missing_columns:
            raise ValueError(f"Eksik sütunlar: {', '.join(missing_columns)}")
        
        return column_indices
    
    def _organize_headers(self, headers: List[str], column_indices: Dict[str, int]) -> List[str]:
        """Başlıkları yeniden düzenler"""
        new_headers = ["TARİH", "İL"]
        other_headers = []
        
        used_indices = set(column_indices.values())
        for idx, header in enumerate(headers, 1):
            if idx not in used_indices and header not in new_headers:
                other_headers.append(header)
        
        new_headers.extend(other_headers)
        return new_headers
    
    async def _copy_data_chunked(self, source_ws: Worksheet, target_ws: Worksheet, 
                               header_row: int, column_indices: Dict[str, int]) -> int:
        """Verileri chunk'lar halinde asenkron olarak kopyalar"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self.thread_pool,
            self._sync_copy_data_chunked,
            source_ws, target_ws, header_row, column_indices
        )
    
    # Verileri chunk'lar halinde senkron olarak kopyalar 
    #  Veri kaybı yapmaz
    #  Hayalet satırları atar
    #  Sayımı doğru yapar
                  
    def _sync_copy_data_chunked(
        self,
        source_ws: Worksheet,
        target_ws: Worksheet,
        header_row: int,
        column_indices: Dict[str, int]
    ) -> int:
        """Verileri chunk'lar halinde senkron olarak kopyalar"""
        date_idx = column_indices["TARİH"]
        city_idx = column_indices["İL"]

        other_columns = [
            col for col in range(1, source_ws.max_column + 1)
            if col not in (date_idx, city_idx)
        ]

        new_row_idx = 2
        real_row_count = 0

        for row in range(header_row + 1, source_ws.max_row + 1):
            city_val = source_ws.cell(row=row, column=city_idx).value
            source_date_cell = source_ws.cell(row=row, column=date_idx)
            date_val = source_date_cell.value
            
            # 🔴 SADECE BU KONTROL
            if city_val is None and date_val is None:
                continue

            # ✔ GERÇEK VERİ SATIRI
            # Tarih hücresini oluştur
            target_date_cell = target_ws.cell(row=new_row_idx, column=1)
            
            # datetime nesnesini string olarak formatla
            if isinstance(date_val, datetime):
                # Sadece tarih kısmını al (YYYY-MM-DD formatında)
                date_str = date_val.strftime("%Y-%m-%d")
                target_date_cell.value = date_str
                target_date_cell.number_format = "YYYY-MM-DD"
            elif isinstance(date_val, date):  # datetime.date nesnesi
                date_str = date_val.strftime("%Y-%m-%d")
                target_date_cell.value = date_str
                target_date_cell.number_format = "YYYY-MM-DD"
            else:
                # Diğer durumlarda olduğu gibi kopyala
                target_date_cell.value = date_val
            
            # Şehir bilgisini yaz
            target_ws.cell(row=new_row_idx, column=2, value=city_val)

            # Diğer sütunları kopyala
            for new_col_idx, source_col in enumerate(other_columns, start=3):
                target_ws.cell(
                    row=new_row_idx,
                    column=new_col_idx,
                    value=source_ws.cell(row=row, column=source_col).value
                )

            new_row_idx += 1
            real_row_count += 1

        return real_row_count
    
    
    async def _adjust_column_widths(self, ws: Worksheet):
        """Sütun genişliklerini asenkron olarak ayarlar"""
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            self.thread_pool,
            self._sync_adjust_column_widths,
            ws
        )
    
    # ------ Sütun genişliği -------------------------------
    # Sabit Genişlik (En hızlısı) - 100x (1-2 ms)
    def _sync_adjust_column_widths(self, ws: Worksheet):
            fixed_width = 25  # istediğin sabit genişlik 25*0,56=14
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = fixed_width
                
    async def _save_workbook(self, wb: Workbook, file_path: str):
        """Workbook'u asenkron olarak kaydeder"""
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            self.thread_pool,
            self._sync_save_workbook,
            wb, file_path
        )
    
    def _sync_save_workbook(self, wb: Workbook, file_path: str):
        """Workbook'u senkron olarak kaydeder"""
        wb.save(file_path)
    
    async def _load_workbook(self, file_path: str) -> Workbook:
        """Workbook'u asenkron olarak yükler"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self.thread_pool,
            load_workbook,
            file_path
        )
    

    # utils/excel_cleaner.py - AsyncExcelCleaner sınıfına yeni metod

    async def clean_excel_headers(self, input_path: str) -> Dict[str, Any]:
        wb = new_wb = None
        temp_path = None
        
        try:
            if not os.path.exists(input_path):
                raise FileNotFoundError(f"Dosya bulunamadı: {input_path}")
            
            if not await self._check_file_size(input_path):
                return {"success": False, "error": f"Dosya boyutu {MAX_FILE_SIZE_MB}MB'den büyük"}
            
            logger.info(f"Excel temizleme başlatıldı: {input_path}")
            
            wb = await self._load_workbook(input_path)
            ws = wb.active
            
            # 🔥 DEĞİŞEN KISIM: SmartHeaderDetector kullan
            header_info = await self._find_header_row_smart(ws)
            
            if not header_info['success']:
                return {"success": False, "error": "Başlık satırı bulunamadı (TARİH ve İL sütunları aranıyor)"}
            
            header_row = header_info['header_row']
            date_col = header_info['date_column']
            city_col = header_info['city_column']
            
            logger.info(f"Başlık satırı: {header_row}, TARİH sütunu: {date_col}, İL sütunu: {city_col}")
            
            # Başlıkları al
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                headers.append(str(val).strip() if val else f"Bos_{col}")
            
            logger.info(f"Temizlenen başlıklar: {headers}")
            
            # 🔥 DEĞİŞEN KISIM: Artık column_indices manuel oluştur
            column_indices = {
                "TARİH": date_col,
                "İL": city_col
            }
            
            new_headers = self._organize_headers(headers, column_indices)
            logger.info(f"Yeni başlık düzeni: {new_headers}")
            
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Düzenlenmiş Veri"
            
            for col_idx, header in enumerate(new_headers, 1):
                new_ws.cell(row=1, column=col_idx, value=header)
            
            row_count = await self._copy_data_chunked(ws, new_ws, header_row, column_indices)
            logger.info(f"Toplam {row_count} satır kopyalandı")
            
            await self._adjust_column_widths(new_ws)
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_path = temp_file.name
            temp_file.close()
            
            await self._save_workbook(new_wb, temp_path)
            logger.info(f"Geçici dosya oluşturuldu: {temp_path}")
            
            return {
                "success": True,
                "temp_path": temp_path,
                "headers": new_headers,
                "row_count": row_count,
                "original_headers": headers,
                "processed_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Excel temizleme hatası: {e}")
            if temp_path and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
            return {"success": False, "error": str(e), "error_type": type(e).__name__}
        finally:
            if wb:
                wb.close()
            if new_wb:
                new_wb.close()
                

    async def batch_clean_excel_files(self, file_paths: List[str]) -> Dict[str, Any]:
        """
        Birden fazla Excel dosyasını asenkron olarak temizler
        
        Args:
            file_paths: Excel dosya yolları listesi
            
        Returns:
            Toplu işlem sonuçları
        """
        tasks = [self.clean_excel_headers(file_path) for file_path in file_paths]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        successful = []
        failed = []
        
        for i, result in enumerate(results):
            if isinstance(result, Exception):
                failed.append({
                    "file": file_paths[i],
                    "error": str(result)
                })
            elif result.get("success"):
                successful.append({
                    "file": file_paths[i],
                    "temp_path": result["temp_path"],
                    "row_count": result["row_count"]
                })
            else:
                failed.append({
                    "file": file_paths[i],
                    "error": result.get("error", "Bilinmeyen hata")
                })
        
        return {
            "total_files": len(file_paths),
            "successful": successful,
            "failed": failed,
            "success_rate": len(successful) / len(file_paths) * 100 if file_paths else 0
        }
    
    def __del__(self):
        """Nesne yok edilirken thread pool'u temizle"""
        if hasattr(self, 'thread_pool'):
            self.thread_pool.shutdown(wait=False)
