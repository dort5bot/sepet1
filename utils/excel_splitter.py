# utils/excel_splitter.py - GÜNCELLENMİŞ
import asyncio
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Set
import functools
from concurrent.futures import ThreadPoolExecutor

import xlsxwriter
from openpyxl import load_workbook

from utils.group_manager import group_manager
from utils.file_namer import generate_output_filename
from utils.logger import logger
from config import config

# Worker pool for sync IO (openpyxl + xlsxwriter)
_DEFAULT_POOL = ThreadPoolExecutor(max_workers=4)


def _sync_read_all_rows(path: str) -> List[tuple]:
    """Senkron: workbook'u açıp tüm satırları (values_only) okur ve kapatır."""
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()
    return rows


def _sync_create_writer(file_path: Path, headers: List[str], sheet_name: str = "Veriler") -> Tuple[xlsxwriter.Workbook, xlsxwriter.worksheet.Worksheet]:
    wb = xlsxwriter.Workbook(file_path, {'constant_memory': True})
    ws = wb.add_worksheet(sheet_name)
    ws.write_row(0, 0, headers)
    ws.set_column(0, len(headers) - 1, 15)
    return wb, ws


def _sync_close_writer(wb: xlsxwriter.Workbook):
    try:
        wb.close()
    except Exception:
        raise


def _sync_write_row(ws: xlsxwriter.worksheet.Worksheet, row_index: int, row: tuple):
    ws.write_row(row_index, 0, row)


class ExcelSplitter:
    """
    Async-friendly ExcelSplitter:
    - Senkron heavy IO'yu threadpool'a atar (openpyxl, xlsxwriter)
    - group_manager çağrılarını şehir bazlı cache'ler
    """

    def __init__(self, input_path: str, headers: List[str], executor: ThreadPoolExecutor = None):
        self.input_path = input_path
        self.headers = headers
        self.writers: Dict[str, xlsxwriter.Workbook] = {}
        self.sheets: Dict[str, Any] = {}
        self.row_counts: Dict[str, int] = {}
        self.matched_row_ids: Set[int] = set()
        self.unmatched_data: List[tuple] = []
        self.unmatched_cities: Set[str] = set()
        self._city_cache: Dict[Any, List[str]] = {}
        self._executor = executor or _DEFAULT_POOL
        
        # YENİ: Gruplara göre şehir bilgisi
        self.group_cities: Dict[str, Set[str]] = {}

    # ---------- helpers ----------
    async def _read_rows(self) -> List[tuple]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(self._executor, functools.partial(_sync_read_all_rows, self.input_path))

    async def _ensure_group_writer(self, group_id: str) -> None:
        """Create workbook + sheet for group if not exists (in threadpool)."""
        if group_id in self.writers:
            return

        group_info = await group_manager.get_group_info(group_id)
        filename = await generate_output_filename(group_info)

        output_dir = config.paths.OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / filename

        loop = asyncio.get_running_loop()
        wb_ws = await loop.run_in_executor(self._executor, functools.partial(_sync_create_writer, file_path, self.headers))
        wb, ws = wb_ws
        self.writers[group_id] = wb
        self.sheets[group_id] = ws
        self.row_counts[group_id] = 1
        
        # YENİ: Grup için şehir seti oluştur
        self.group_cities[group_id] = set()
        
        logger.debug(f"excelsplit Writer created for group {group_id}: {file_path}")

    async def _write_row(self, group_id: str, row: tuple) -> None:
        """Write a row to group's sheet (performed in threadpool)."""
        ws = self.sheets[group_id]
        row_index = self.row_counts[group_id]
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(self._executor, functools.partial(_sync_write_row, ws, row_index, row))
        self.row_counts[group_id] += 1

    async def _close_all_writers(self) -> Dict[str, Dict[str, Any]]:
        """Close workbooks in threadpool and return output_files dict."""
        output_files: Dict[str, Dict[str, Any]] = {}
        loop = asyncio.get_running_loop()

        for group_id, wb in list(self.writers.items()):
            try:
                row_count = self.row_counts.get(group_id, 1) - 1
                await loop.run_in_executor(self._executor, functools.partial(_sync_close_writer, wb))
                ws_path = Path(wb.filename)
                
                if row_count <= 0:
                    if ws_path.exists():
                        try:
                            ws_path.unlink()
                            logger.info(f"🗑️ excelsplit Deleted empty file: {ws_path.name}")
                        except Exception as e:
                            logger.warning(f"excelsplit Failed to delete empty file {ws_path}: {e}")
                    continue

                # YENİ: Şehir bilgisini output_files'a ekle
                cities_in_group = list(self.group_cities.get(group_id, set()))
                
                output_files[group_id] = {
                    "filename": ws_path.name,
                    "path": ws_path,
                    "row_count": row_count,
                    "cities": cities_in_group  # YENİ!
                }
                
                logger.info(f"📄 Saved: {ws_path.name} ({row_count} rows, {len(cities_in_group)} cities)")
                
            except Exception as e:
                logger.error(f"excelsplit Error closing workbook for {group_id}: {e}", exc_info=True)

        return output_files

    # ---------- main ----------
    async def run(self) -> Dict[str, Any]:
        try:
            logger.info("🔄 excelsplit Group manager initializing…")
            await group_manager._ensure_initialized()

            logger.info("📥 excelsplit Reading input file…")
            rows = await self._read_rows()
            processed_rows = 0

            # Build unique city set for caching
            cities = set()
            for row in rows:
                if row and len(row) > 1 and row[1]:
                    cities.add(row[1])

            logger.debug(f"excelsplit Unique cities found: {len(cities)}")

            # Fetch groups for each city concurrently and cache results
            city_tasks = {city: asyncio.create_task(group_manager.get_groups_for_city(city)) 
                         for city in cities}
            
            for city, task in city_tasks.items():
                try:
                    groups = await task
                    self._city_cache[city] = groups or []
                except Exception as e:
                    logger.warning(f"excelsplit City lookup failed for {city}: {e}")
                    self._city_cache[city] = []

            # Process rows
            for row_idx, row in enumerate(rows):
                processed_rows += 1
                city = row[1] if row and len(row) > 1 else None
                groups = []
                
                if city:
                    groups = self._city_cache.get(city, [])

                has_match = False
                if groups:
                    for g in groups:
                        if g != "grup_0":
                            has_match = True
                            self.matched_row_ids.add(row_idx)
                            
                            await self._ensure_group_writer(g)
                            await self._write_row(g, row)
                            
                            # YENİ: Şehri gruba ekle
                            if city and g in self.group_cities:
                                self.group_cities[g].add(city)

                if not has_match:
                    if city:
                        self.unmatched_cities.add(city)
                        self.unmatched_data.append(row)
            
            logger.info(f"✔ excelsplit Processing complete. Total rows processed: {processed_rows}")

            # Finalize writers
            output_files = await self._close_all_writers()

            # If unmatched data exists, create grup_0 file
            if self.unmatched_data:
                group_id = "grup_0"
                try:
                    group_info = await group_manager.get_group_info(group_id)
                    filename = await generate_output_filename(group_info)
                    output_dir = config.paths.OUTPUT_DIR
                    output_dir.mkdir(parents=True, exist_ok=True)
                    file_path = output_dir / filename

                    loop = asyncio.get_running_loop()
                    wb_ws = await loop.run_in_executor(self._executor, 
                        functools.partial(_sync_create_writer, file_path, self.headers, "Eşleşmeyenler"))
                    wb, ws = wb_ws
                    
                    for idx, row in enumerate(self.unmatched_data, start=1):
                        await loop.run_in_executor(self._executor, 
                            functools.partial(_sync_write_row, ws, idx, row))
                    
                    await loop.run_in_executor(self._executor, 
                        functools.partial(_sync_close_writer, wb))

                    output_files[group_id] = {
                        "filename": file_path.name,
                        "path": file_path,
                        "row_count": len(self.unmatched_data),
                        "cities": list(self.unmatched_cities)  # YENİ: Eşleşmeyen şehirler
                    }
                    
                    logger.info(f"📄 excelsplit Eşleşmeyenler dosyası: {filename} ({len(self.unmatched_data)} satır)")
                    
                except Exception as e:
                    logger.error(f"excelsplit Eşleşmeyenler dosyası hatası: {e}", exc_info=True)

            return {
                "success": True,
                "processed_rows": processed_rows,
                "matched_rows": len(self.matched_row_ids),
                "unmatched_rows": len(self.unmatched_data),
                "output_files": output_files,
                "unmatched_cities": list(self.unmatched_cities),
            }

        except Exception as e:
            logger.error(f"❌ Error in ExcelSplitter.run: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "processed_rows": 0,
                "output_files": {},
            }

# external API
async def split_excel_by_groups(input_path: str, headers: List[str]) -> Dict[str, Any]:
    splitter = ExcelSplitter(input_path, headers)
    return await splitter.run()


def split_excel_by_groups_sync(input_path: str, headers: List[str]) -> Dict[str, Any]:
    return asyncio.run(split_excel_by_groups(input_path, headers))