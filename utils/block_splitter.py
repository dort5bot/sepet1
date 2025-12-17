
# utils/block_splitter.py
"""
Blok bazlı Excel bölücü – PERFORMANS ODAKLI SÜRÜM

Değişiklik özeti:
- openpyxl iter_rows(values_only=True)
- Tek geçiş: blok tespiti + kopyalama
- ThreadPool sadece workbook açma / kapama için
- Satır yazımı senkron (xlsxwriter zaten buffer'lı)
- Hücre hücre erişim YOK
- pandas YOK (gereksiz)
"""

import unicodedata
import asyncio
from pathlib import Path
from typing import Dict, List, Any, Tuple
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor

import xlsxwriter
from openpyxl import load_workbook

from utils.group_manager import group_manager
from utils.file_namer import generate_output_filename
from utils.logger import logger
from config import config


class BlockExcelSplitter:
    _CITY_SET: set[str] | None = None

    def __init__(self, input_path: str, headers: List[str]):
        self.input_path = input_path
        self.headers = headers
        self._executor = ThreadPoolExecutor(max_workers=2)

    # --------------------------------------------------
    # Normalize
    # --------------------------------------------------
    @staticmethod
    def normalize_turkish(text) -> str:
        if not text or not isinstance(text, str):
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = text.replace("\u0307", "")
        return text.strip().upper()

    # --------------------------------------------------
    # City set (cache)
    # --------------------------------------------------
    @classmethod
    def _get_city_set(cls) -> set[str]:
        if cls._CITY_SET is None:
            raw = [
                "Adana","Adıyaman","Afyon","Ağrı","Amasya","Ankara",
                "Antalya","Artvin","Aydın","Balıkesir","Bilecik","Bingöl",
                "Bitlis","Bolu","Burdur","Bursa","Çanakkale","Çankırı",
                "Çorum","Denizli","Diyarbakır","Edirne","Elazığ","Erzincan",
                "Erzurum","Eskişehir","Gaziantep","Giresun","Gümüşhane",
                "Hakkari","Hatay","Isparta","İçel","İstanbul","İzmir",
                "Kars","Kastamonu","Kayseri","Kırklareli","Kırşehir",
                "Kocaeli","Konya","Kütahya","Malatya","Manisa",
                "Kahramanmaraş","Mardin","Muğla","Muş","Nevşehir",
                "Niğde","Ordu","Rize","Sakarya","Samsun","Siirt",
                "Sinop","Sivas","Tekirdağ","Tokat","Trabzon","Tunceli",
                "Şanlıurfa","Uşak","Van","Yozgat","Zonguldak","Aksaray",
                "Bayburt","Karaman","Kırıkkale","Batman","Şırnak",
                "Bartın","Ardahan","Iğdır","Yalova","Karabük",
                "Kilis","Osmaniye","Düzce"
            ]
            cls._CITY_SET = {cls.normalize_turkish(c) for c in raw}
        return cls._CITY_SET

    # --------------------------------------------------
    # City row check
    # --------------------------------------------------
    def _is_city_row(self, cell_value) -> tuple[bool, str]:
        value = self.normalize_turkish(cell_value)
        if not value:
            return False, ""

        city_set = self._get_city_set()

        if value in city_set:
            return True, value

        if value.endswith(" İLİ"):
            city = value[:-4]
        elif value.endswith(" İL"):
            city = value[:-3]
        else:
            return False, ""

        return (city in city_set), city if city in city_set else ""

    # --------------------------------------------------
    # MAIN PROCESS
    # --------------------------------------------------
    async def process_blocks(self) -> Dict[str, Any]:
        try:
            await group_manager._ensure_initialized()
            loop = asyncio.get_running_loop()

            # Workbook load (executor)
            wb = await loop.run_in_executor(
                self._executor,
                lambda: load_workbook(self.input_path, read_only=True)
            )
            ws = wb.active

            output_files: Dict[str, Dict] = {}
            total_rows = 0
            current_city: str | None = None

            logger.info(f"📊 Excel satır sayısı: {ws.max_row}")

            # --------------------------------------------------
            # SINGLE PASS
            # --------------------------------------------------
            for idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                city_cell = row[1] if len(row) > 1 else None
                is_city, city_name = self._is_city_row(city_cell)

                if is_city:
                    current_city = city_name
                    logger.info(f"📍 Şehir bulundu: {city_name} (satır {idx})")
                    continue

                if not current_city:
                    continue

                # Satırda veri var mı?
                if not any(row):
                    continue

                # Şehrin grupları
                group_ids = await group_manager.get_groups_for_city(current_city)
                if not group_ids:
                    group_ids = ["grup_0"]

                for group_id in group_ids:
                    if group_id not in output_files:
                        group_info = await group_manager.get_group_info(group_id)
                        filename = await generate_output_filename(group_info)

                        out_dir = config.paths.OUTPUT_DIR
                        out_dir.mkdir(parents=True, exist_ok=True)
                        file_path = out_dir / filename

                        wb_out = xlsxwriter.Workbook(
                            file_path, {'constant_memory': True}
                        )
                        ws_out = wb_out.add_worksheet("Veriler")
                        ws_out.write_row(0, 0, self.headers)
                        ws_out.set_column(0, len(self.headers) - 1, 15)

                        output_files[group_id] = {
                            "writer": wb_out,
                            "worksheet": ws_out,
                            "row_count": 1,
                            "file_path": file_path,
                            "cities": set()
                        }

                    out = output_files[group_id]

                    # Row normalize
                    row_data = list(row)
                    if len(row_data) > 1:
                        row_data[1] = current_city

                    for i, v in enumerate(row_data):
                        if isinstance(v, (datetime, date)):
                            row_data[i] = v.strftime("%d.%m.%Y")

                    out["worksheet"].write_row(
                        out["row_count"], 0, row_data
                    )
                    out["row_count"] += 1
                    out["cities"].add(current_city)
                    total_rows += 1

            # --------------------------------------------------
            # CLOSE FILES
            # --------------------------------------------------
            final_outputs = {}
            for gid, data in output_files.items():
                data["writer"].close()

                if data["row_count"] <= 1:
                    try:
                        data["file_path"].unlink()
                    except:
                        pass
                    continue

                final_outputs[gid] = {
                    "filename": data["file_path"].name,
                    "path": data["file_path"],
                    "row_count": data["row_count"] - 1,
                    "cities": list(data["cities"])
                }

                logger.info(
                    f"📄 {data['file_path'].name}: {data['row_count']-1} satır"
                )

            wb.close()

            return {
                "success": True,
                "total_rows": total_rows,
                "total_files": len(final_outputs),
                "output_files": final_outputs
            }

        except Exception as e:
            logger.error(f"❌ Blok işleme hatası: {e}", exc_info=True)
            return {"success": False, "error": str(e)}


# --------------------------------------------------
# PUBLIC API
# --------------------------------------------------
async def split_excel_by_blocks(input_path: str, headers: List[str]) -> Dict[str, Any]:
    splitter = BlockExcelSplitter(input_path, headers)
    return await splitter.process_blocks()
