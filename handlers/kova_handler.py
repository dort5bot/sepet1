# handlers/kova_handler.py
"""
17-11-2025

Upload Handler Module
Excel dosya yükleme ve işleme işlemleri
"""
from pathlib import Path
from typing import Dict, Any
from aiogram import Router, F
import traceback

from aiogram.filters import Command, CommandObject
from aiogram.types import Message

from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from openpyxl import load_workbook

from config import config
from utils.excel_process import process_excel_task
from utils.reporter import generate_processing_report
from utils.logger import logger

# Handler loader uyumlu router tanımı
router = Router(name="kova_processor")

class ProcessingStates(StatesGroup):
    """Dosya işleme state'leri"""
    waiting_for_file = State()

# Sabitler
CANCEL_COMMANDS = {"cancel", "iptal", "stop", "dur", "🛑 dur"}
# EXCEL_EXTENSIONS = {'.xlsx', '.xls'}
EXCEL_EXTENSIONS = {ext.lower() for ext in [".xlsx", ".xls"]}

REQUIRED_COLUMNS = {"TARİH", "İL"}

def _validate_excel_file(file_path: Path) -> Dict[str, Any]:
    """
    Excel dosyasını doğrular
    """
    wb = None
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        
        # Başlık satırını al
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value).strip().upper() if cell_value else "")
        
        # Gerekli sütunları kontrol et
        found_columns = set(headers)
        
        if not REQUIRED_COLUMNS.issubset(found_columns):
            missing = REQUIRED_COLUMNS - found_columns
            return {
                "valid": False,
                "message": f"Dosyada gerekli sütunlar bulunamadı: {', '.join(missing)}"
            }
        
        # Satır sayısını kontrol et (sadece başlık varsa)
        if ws.max_row <= 1:
            return {
                "valid": False,
                "message": "Dosyada işlenecek veri bulunamadı"
            }
        
        return {
            "valid": True, 
            "headers": headers, 
            "row_count": ws.max_row - 1
        }
        
    except Exception as e:
        return {
            "valid": False, 
            "message": f"Dosya okunamadı: {str(e)}"
        }
    finally:
        if wb:
            wb.close()

async def _download_user_file(bot, file_id: str, file_name: str) -> Path:
    """
    Kullanıcı dosyasını indirir
    """
    try:
        file_info = await bot.get_file(file_id)
        file_path = config.paths.INPUT_DIR / file_name  
        await bot.download_file(file_info.file_path, file_path)
        logger.info(f"Dosya indirildi: {file_path}")
        return file_path
    except Exception as e:
        logger.error(f"Dosya indirme hatası: {e}")
        raise


async def _process_uploaded_file(message: Message, file_path: Path) -> Dict[str, Any]:
    """
    Yüklenen dosyayı işler
    """
    try:
        logger.info(f"Dosya işleniyor: {file_path}")
        
        # Doğrulama
        validation_result = _validate_excel_file(file_path)
        if not validation_result["valid"]:
            logger.error(f"Doğrulama hatası: {validation_result['message']}")
            return {
                "success": False, 
                "error": validation_result["message"]
            }
        
        logger.info(f"Doğrulama başarılı: {validation_result['row_count']} satır")
        
        # İşlemi başlat
        task_result = await process_excel_task(file_path, message.from_user.id)
        logger.info(f"İşlem sonucu: {task_result}")
        return task_result
        
    except Exception as e:
        error_msg = f"İşlem sırasında hata oluştu: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return {
            "success": False, 
            "error": error_msg
        }

@router.message(Command("start"))
async def cmd_start(message: Message):
    """
    /start komutu - hoşgeldin mesajı
    """
    await message.answer(
        "📊 Excel İşleme Botuna Hoşgeldiniz! - kova\n\n"
        #"Lütfen işlemek istediğiniz Excel dosyasını gönderin.\n"
        #"Dosyada 1.satırda 'TARİH' ve 'İL' sütunları bulunmalıdır.\n"
        " Tüm işlemleri görmek için sanal klavyeye geç\n"
        "sanal klavye için tıkla  yada yaz:  /r  \n"
        "sonra açıklamaları görmek için 'oku' butonuna bas"
    )

# /process = /kova (aynı iş)
@router.message(Command("kova", "process"))  # ✅ İki komut tek handler
async def cmd_process(message: Message, state: FSMContext):
    """
    /process VE /kova komutları - aynı işi yapar
    """
    await state.set_state(ProcessingStates.waiting_for_file)
    await message.answer(
        "📌 (Zorunlu) 1.satırda İL, TARİH yazılmalıdır\n"
        "📤 İşlemek istediğin Excel dosyasını gönder...\n"
        "🛑 İptal için tıkla: '/iptal' veya bas: DUR"
    )

    
    

@router.message(ProcessingStates.waiting_for_file, F.text)
async def handle_cancel_command(message: Message, state: FSMContext):
    """
    İptal komutlarını yakalar
    """
    user_text = message.text.strip().lower().lstrip('/')
    
    if user_text in CANCEL_COMMANDS:
        await state.clear()
        await message.answer(
            "❌ İşlem iptal edildi.\n"
            "Ana menüye dönmek için /start komutunu kullanabilirsiniz."
        )
    else:
        await message.answer(
            "❌ Lütfen bir Excel dosyası gönderin veya /iptal komutu ile işlemi iptal edin."
        )

# Excel dosyası yükleme handler

@router.message(ProcessingStates.waiting_for_file, F.document)
async def handle_excel_upload(message: Message, state: FSMContext):
    """
    Excel dosyası yükleme handler'ı - UZUN VADELİ FİNAL VERSİYON
    """

    original_name = message.document.file_name

    # Uzantıyı küçült
    file_ext = Path(original_name).suffix.lower()

    # Uzantı kontrolü (kullanıcı .XLS bile gönderse çalışır)
    if file_ext not in EXCEL_EXTENSIONS:
        await message.answer("❌ Lütfen Excel dosyası (.xlsx veya .xls) gönderin.")
        await state.clear()
        return

    # Dosya adını normalize et (KULLANICI NE YAZARSA YAZSIN temiz format)
    clean_name = Path(original_name).stem
    normalized_name = f"{clean_name}{file_ext}"   # Örn: RAPOR.XLSX → RAPOR.xlsx

    file_path = None

    try:
        logger.info(f"Dosya alındı: {original_name}, Boyut: {message.document.file_size}")

        # 1. Dosyayı normalize edilmiş isimle indir
        await message.answer("📥 Dosya indiriliyor...")
        file_path = await _download_user_file(
            message.bot,
            message.document.file_id,
            normalized_name  # ← ARTIK HER ZAMAN KÜÇÜK HARFLİ UZANTI
        )
        logger.info(f"Dosya indirme tamamlandı: {file_path}")

        # 2. Doğrulama
        await message.answer("🔍 Dosya kontrol ediliyor...")
        validation_result = _validate_excel_file(file_path)
        if not validation_result["valid"]:
            await message.answer(f"❌ {validation_result['message']}")
            await state.clear()
            return

        logger.info(f"Doğrulama başarılı: {validation_result['row_count']} satır")

        # 3. Dosya işleme
        await message.answer("⏳ Dosya işleniyor, lütfen bekleyin...")
        task_result = await _process_uploaded_file(message, file_path)

        if task_result["success"]:
            # report = await generate_processing_report(task_result)
            # report = generate_processing_report(task_result)
            
            report = generate_processing_report(
                task_result,
                for_internal_message=True
            )
                  
                        
            
            
            await message.answer(report)
            logger.info("İşlem başarıyla tamamlandı")
        else:
            error_msg = f"❌ {task_result['error']}"
            if len(error_msg) > 1000:
                error_msg = error_msg[:1000] + "... (devamı loglarda)"
            await message.answer(error_msg)
            logger.error(f"İşlem hatası: {task_result['error']}")

    except Exception as e:
        error_msg = str(e)
        if len(error_msg) > 500:
            error_msg = error_msg[:500] + "..."
        await message.answer(f"❌ İşlem sırasında hata oluştu: {error_msg}")
        logger.error(traceback.format_exc())

    finally:
        if file_path and file_path.exists():
            try:
                file_path.unlink()
                logger.info(f"Geçici dosya silindi: {file_path}")
            except Exception as e:
                logger.warning(f"Dosya silinemedi {file_path}: {e}")
        await state.clear()



@router.message(ProcessingStates.waiting_for_file)
async def handle_wrong_file_type(message: Message):
    """
    Yanlış dosya tipi handler'ı
    """
    await message.answer("❌ Lütfen bir Excel dosyası gönderin.")